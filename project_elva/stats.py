import json
import re
from datetime import datetime
from tqdm import *

# ---------------------------------
# DB config
# ---------------------------------

from models import (Base,
        Computer, Instance,
        School, Session,
        Student, Teacher,
        Unit
)
from sqlalchemy import create_engine
from sqlalchemy import MetaData
from sqlalchemy.engine.url import URL
from sqlalchemy.orm import sessionmaker

db_url = {
        'drivername': 'postgresql',
        'database': os.environ['POSTGRES_DB'], 
        'username': os.environ['POSTGRES_USER'],
        'password': os.environ['POSTGRES_PASSWORD'],
        'host': 'localhost',
        'port': 5432
}
engine = create_engine(URL(**db_url))

DBSession = sessionmaker(bind=engine)
session = DBSession()

# ---------------------------------
# XLSX output config 
# ---------------------------------

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border,
                            Font, PatternFill,
                            Side)
from openpyxl.utils import (coordinate_from_string, column_index_from_string)

NULL_TOLERANCE = 50
DATESTRING = str(datetime.utcnow())[:-7].\
                 replace(':', '-').\
                 replace(' ', '_')

class ExcelWriter(object):

    def __init__(self, DBSession, workbook=Workbook()):
        self.__computer_row_count = 0
        self.__log_dict = None
        self.__school_row_count = 0
        self._session = DBSession
        self.__student_row_count = 0
        self.__teacher_row_count = 0
        self._workbook = workbook

    @property
    def workbook(self):
        return self._workbook

    @workbook.setter
    def workbook(self, workbook=Workbook()):
        self._workbook = workbook

    def add_computer(self, worksheet, school_name, teacher, computer):
        """
        Add a computer to the spreadsheet, and then recursively add
        the students.

        """
        
        alignment = Alignment(horizontal='center',
                              vertical='center')
        border = Border(left=Side(border_style='medium',
                                  color='000000'),
                        right=Side(border_style='medium',
                                   color='000000'),
                        top=Side(border_style='medium',
                                 color='000000'),
                        bottom=Side(border_style='medium',
                                    color='000000'))
        font = Font(name='Calibri',
                    size=12,
                    color='000000')
       
        num_students = int(self.__log_dict[school_name][teacher][computer]['student_count'])
        self.style_range(worksheet,
                         3, self.__computer_row_count+3,
                         3, self.__computer_row_count+(3*num_students)+2,
                         alignment=alignment, border=border,
                         font=font, value=computer)

        self.__computer_row_count += num_students * 3

        students = self.__log_dict[school_name][teacher][computer].keys()
        for student in students:
            if student != 'student_count':
                self.add_student(worksheet, school_name, teacher, computer, student)

    def add_legend(self, title='Legend'):
        """Add a legend sheet to the spreadsheet."""
        
        # styling options
        alignment = Alignment(horizontal='center',
                              vertical='center')
        border = Border(left=Side(border_style='thin',
                                  color='000000'),
                        right=Side(border_style='thin',
                                   color='000000'),
                        bottom=Side(border_style='medium',
                                    color='000000'))
        bottom = Border(bottom=Side(border_style='thin',
                                    color='000000'))
        fill = PatternFill(fill_type='solid',
                           start_color='d9d9d9',
                           end_color='d9d9d9')
        font = Font(name='Calibri',
                    size=12,
                    bold=True,
                    color='000000')
        top = Border(top=Side(border_style='thin',
                              color='000000'))
        
        ws2 = self._workbook.create_sheet(title)

        data = {'student_id': 103065,
                'session': 'CesarChavez_1',
                'time': '0:07:03',
                'slides': 63,
                'complete': 23}

        # add the data to the legend
        ws2['A1'].value = 'Student ID'
        ws2['A2'].value = 'Student'
        self.style_range(ws2, 1, 4, 1, 6,
                         alignment=alignment, value=data['student_id'])

        ws2['A8'].value = 'NOTE: Each session has five columns to accomodate situations where a student has started a session more than once.'

        ws2['B1'].value = 'Session'
        ws2['B2'].value = data['session']
        ws2['B3'].value = data['slides']
        ws2['B4'].value = data['time']
        ws2['B5'].value = data['complete']
        ws2['B6'].value = float(data['complete']) / float(data['slides'])

        ws2['C3'].value = '# of slides contained in a session.'
        ws2['C4'].value = 'time logged for an instance of a session.'
        ws2['C5'].value = 'number of slides a student has completed during an instance of a session.'
        ws2['C6'].value = 'percent complete (based on slides complete divided by slides total).'

        # add the styling to the legend
        ws2['A2'].alignment = alignment
        ws2['A2'].border = ws2['A2'].border + border
        ws2['A2'].fill = fill
        ws2['A2'].font = font

        ws2['B2'].alignment = alignment
        ws2['B2'].border = ws2['B2'].border + border
        ws2['B2'].fill = fill
        ws2['B2'].font = font

        ws2['B4'].alignment = alignment
        ws2['B4'].border = ws2['B4'].border + top
        
        ws2['B5'].alignment = alignment

        ws2['B6'].alignment = alignment
        ws2['B6'].border = ws2['B6'].border + bottom

        # changelog values
        ws2['A10'].value = 'Changelog'
        ws2['A10'].font = font

        strings = ['', 
                   '04/24/2017 - Added freeze panes to make it easier to track session information and student information.',
                   '04/18/2017 - Added sorting and colors to highlight the first time a student attempted a session.',
                   '04/17/2017 - Added functionality to widen columns.',
                   '']

        index = 1
        for row in ws2.iter_rows(min_col=1,
                                 max_col=1,
                                 min_row=11,
                                 max_row=11+(len(strings)-3)): 
            if len(strings) > 2:
                cell = row[0]
                cell.value = strings[index]
                index += 1


    def add_log_data(self, worksheet, log_dict={}):
        """
        Add the log data dict to the object and recursively add the data
        into the worksheet.

        """

        if len(log_dict.keys()) == 0:
            return
        else:
            self.__log_dict = log_dict

        for school in tqdm(self.__log_dict.keys()):
            self.add_school(worksheet, school)

    def add_school(self, worksheet, school_name):
        """
        Add a school to the spreadsheet, and then recursively add
        the teachers (etc.).

        """

        alignment = Alignment(horizontal='center',
                              vertical='center')
        border = Border(left=Side(border_style='medium',
                                  color='000000'),
                        right=Side(border_style='medium',
                                   color='000000'),
                        top=Side(border_style='medium',
                                 color='000000'),
                        bottom=Side(border_style='medium',
                                    color='000000'))
        font = Font(name='Calibri',
                    size=12,
                    color='000000')
        num_format = '###'


        num_students = int(self.__log_dict[school_name]['student_count'])
        
        self.style_range(worksheet,
                         1, self.__school_row_count+3,
                         1, self.__school_row_count+(3*num_students)+2,
                         alignment=alignment, border=border,
                         font=font, num_format=num_format,
                         value=school_name)
    
        self.__school_row_count += num_students * 3

        teachers = self.__log_dict[school_name].keys()
        for teacher in teachers:
            if teacher != 'student_count':
                self.add_teacher(worksheet, school_name, teacher)

    def add_student(self, worksheet, school_name, teacher, computer, student):
        """
        Add a student (and the associated data) to the spreadsheet. 

        """

        alignment = Alignment(horizontal='center',
                              vertical='center')
        border = Border(left=Side(border_style='medium',
                                  color='000000'),
                        right=Side(border_style='medium',
                                   color='000000'),
                        top=Side(border_style='medium',
                                 color='000000'),
                        bottom=Side(border_style='medium',
                                    color='000000'))
        fill = PatternFill(fill_type='solid',
                           start_color='9cba5f',
                           end_color='9cba5f')
        font = Font(name='Calibri',
                    size=12,
                    color='000000')
        num_format = '######'

        top_border = Border(top=Side(border_style='thin',
                                     color='000000'))
        bottom_border = Border(bottom=Side(border_style='thin',
                                           color='000000'))
        med_right_border = Border(right=Side(border_style='medium',
                                             color='000000'))
        thin_right_border = Border(right=Side(border_style='thin',
                                              color='000000'))

        self.style_range(worksheet,
                         4, self.__student_row_count+3,
                         4, self.__student_row_count+5,
                         alignment=alignment, border=border,
                         font=font, num_format=num_format,
                         value=student)

        # get instance data for student
        data = self.__log_dict[school_name][teacher][computer][student]      

        # insert the time spent on the intance
        for col in worksheet.iter_cols(min_col=5,
                                       min_row=self.__student_row_count+3,
                                       max_row=self.__student_row_count+3):
            for cell in col:
                column = column_index_from_string(str(cell.column))
                title = worksheet.cell(row=1, column=column)    
                
                if title.value in data.keys():
                    cell.value = data[title.value][0]

                    if title.value[-1] != ')':
                        cell.fill = fill

                cell.alignment = alignment
                cell.border = cell.border + top_border
 
                if title.value[-3:] == '(5)' and \
                   title.value[-4:] != '5(5)':
                    cell.border = cell.border + thin_right_border

                if title.value[-4:] == '5(5)':
                    cell.border = cell.border + med_right_border

        # insert the number of slides completed for the intstance
        for col in worksheet.iter_cols(min_col=5,
                                       min_row=self.__student_row_count+4,
                                       max_row=self.__student_row_count+4):
            for cell in col:
                column = column_index_from_string(str(cell.column))
                title = worksheet.cell(row=1, column=column)

                if title.value in data.keys():
                    cell.value = data[title.value][1]

                    if title.value[-1] != ')':
                        cell.fill = fill

                cell.alignment = alignment

                if title.value[-3:] == '(5)' and \
                   title.value[-4:] != '5(5)':
                    cell.border = cell.border + thin_right_border

                if title.value[-4:] == '5(5)':
                    cell.border = cell.border + med_right_border

        # insert the formula to calculate percent completion for the instance
        for col in worksheet.iter_cols(min_col=5,
                                       min_row=self.__student_row_count+5,
                                       max_row=self.__student_row_count+5):
            for cell in col:
                column = column_index_from_string(str(cell.column))
                title = worksheet.cell(row=1, column=column)
                slide = worksheet.cell(row=2, column=column)
                complete = worksheet.cell(row=cell.row-1, column=column)

                if complete.value or complete.value == 0:
                    worksheet[cell.coordinate] = "="+str(complete.coordinate)+ \
                                                 "/"+str(slide.coordinate)

                    if title.value[-1] != ')':
                        cell.fill = fill

                cell.alignment = alignment
                cell.border = cell.border + bottom_border 

                if title.value[-3:] == '(5)' and \
                   title.value[-4:] != '5(5)':
                    cell.border = cell.border + thin_right_border

                if title.value[-4:] == '5(5)':
                    cell.border = cell.border + med_right_border

        self.__student_row_count += 3

    def add_teacher(self, worksheet, school_name, teacher):
        """
        Add a teacher to the spreadsheet and then recursively add the
        computers (etc.).

        """
        
        alignment = Alignment(horizontal='center',
                              vertical='center')
        border = Border(left=Side(border_style='medium',
                                  color='000000'),
                        right=Side(border_style='medium',
                                   color='000000'),
                        top=Side(border_style='medium',
                                 color='000000'),
                        bottom=Side(border_style='medium',
                                    color='000000'))
        font = Font(name='Calibri',
                    size=12,
                    color='000000')
        num_format = '######'

        num_students = int(self.__log_dict[school_name][teacher]['student_count'])
        self.style_range(worksheet,
                         2, self.__teacher_row_count+3,
                         2, self.__teacher_row_count+(3*num_students)+2,
                         alignment=alignment, border=border,
                         font=font, num_format=num_format,
                         value=teacher)

        self.__teacher_row_count += num_students * 3

        computers = self.__log_dict[school_name][teacher].keys()
        for computer in computers:
            if computer != 'student_count':
                self.add_computer(worksheet, school_name, teacher, computer)

    def create_title_bar(self, worksheet):
        """
        Add a title bar to the worksheet. Contains titles for first four columns
        and session names for the following columns.
    
        """
    
        # styling options
        alignment = Alignment(horizontal='center',
                              vertical='center')
        border = Border(left=Side(border_style='thin',
                                  color='000000'),
                        right=Side(border_style='thin',
                                   color='000000'),
                        bottom=Side(border_style='medium',
                                    color='000000'))
        fill = PatternFill(fill_type='solid',
                           start_color='d9d9d9',
                           end_color='d9d9d9')
        font = Font(name='Calibri',
                    size=12,
                    bold=True,
                    color='000000')
    
        # titles for first four columns
        titles = ['School', 'Teacher', 'Computer', 'Student']
        for n in xrange(1,len(titles)+1):
            cell = worksheet.cell(row=1, column=n, value=titles[n-1])
            
            # apply styles for these cells
            cell.alignment = alignment
            cell.border = cell.border + border
            cell.fill = fill
            cell.font = font
    
        # gather session info from DB
        sessions = self._session.query(Session).order_by(Session.name).all()
    
        # add five more entries for each session
        # to accommodate duplicates
        exp_titles = {}
        for x in xrange(1,6):
            for item in sessions:
                if x == 1:
                    exp_titles[item.name] = item.slides
                exp_titles[item.name+'('+str(x)+')'] = item.slides
        keys = sorted(exp_titles.keys())

        # add sessions to spreadsheet
        following_cell_count = len(sessions) * 6
        current_cell_num = 0
        while current_cell_num < following_cell_count:
            sess = keys[current_cell_num]
            cell = worksheet.cell(row=1,
                                  column=int(current_cell_num+5),
                                  value=sess)

            # add style to top row
            cell.alignment = alignment
            cell.border = cell.border + border
            cell.fill = fill
            cell.font = font
          
            # add the total number of slides to the 2nd row
            cell = worksheet.cell(row=2,
                                  column=int(current_cell_num+5),
                                  value=exp_titles[sess]) 

            current_cell_num += 1

        # adjust width of each column to suit the
        # width of the title value
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0)),
                                            len(str(cell.value)))

        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value

    def save(self, filename='untitled_'+DATESTRING+'.xlsx'):
        """Save the workbook using a given filename."""
        
        self._workbook.save(filename)

    def style_range(self, worksheet,
                    start_col, start_row,
                    end_col, end_row,
                    alignment=None, border=Border(),
                    fill=None, font=None,
                    num_format=None, value=None):
        """Style a range of cells."""
        
        bottom = Border(border.bottom)
        left = Border(border.left)
        right = Border(border.right)
        top = Border(border.top)
        
        first_cell = worksheet.cell(row=start_row, column=start_col)
        if alignment:
            worksheet.merge_cells(start_row=start_row,
                                  start_column=start_col,
                                  end_row=end_row,
                                  end_column=end_col)
            first_cell.alignment = alignment

        if font:
            first_cell.font = font

        if num_format:
            first_cell.number_format = num_format

        if value:
            first_cell.value = value

#        for col in worksheet.iter_cols(min_col=start_col,
#                                       min_row=start_row,
#                                       max_col=end_col,
#                                       max_row=start_row):
#            for cell in col:
#                cell.border = cell.border + top
#
#        for col in worksheet.iter_cols(min_col=start_col,
#                                       min_row=end_row,
#                                       max_col=end_col,
#                                       max_row=end_row):
#            for cell in col:
#                cell.border = cell.border + bottom

        for row in worksheet.iter_rows(min_col=start_col,
                                       min_row=start_row,
                                       max_col=end_col,
                                       max_row=end_row):
            if fill:
                for cell in row:
                    cell.fill = fill

def get_data():
    """
    Return a dictionary containing all of the major datapoints
    (schools, teachers, computers, students, instances, etc.)

    """
    
    global session

    # get all the data you need in one gigantic DB query
    students = session.query(Computer, Instance,
                             School, Session,
                             Student, Teacher).\
                       join(Instance.session).\
                       join(Instance.student).\
                       join(Student.computer).\
                       join(Student.teacher).\
                       join(Teacher.school).\
                       order_by(School.name).\
                       order_by(Teacher.name).\
                       order_by(Computer.guid).\
                       order_by(Student.id).\
                       order_by(Session.name).\
                       order_by(Instance.start_time).\
                       values(School.id, Teacher.id,
                              Computer.guid, Student.id,
                              Session.name, Instance.start_time,
                              Instance.end_time, Instance.slides_finished,
                              Session.slides)

    # collate all of this data into one big dictionary
    sheet = {}
    for school, teacher, \
        guid, id, \
        session_name, start_time, \
        end_time, finished, \
        total in students: 
    
        if school in sheet.keys():
            if teacher in sheet[school].keys():
                if guid in sheet[school][teacher].keys():
                    if id in sheet[school][teacher][guid].keys():
                        pass
                    else:
                        sheet[school][teacher][guid][id] = {}
                        sheet[school]['student_count'] += 1
                        sheet[school][teacher]['student_count'] += 1
                        sheet[school][teacher][guid]['student_count'] += 1
                else:
                    sheet[school][teacher][guid] = {id: {}}
                    sheet[school]['student_count'] += 1
                    sheet[school][teacher]['student_count'] += 1
                    sheet[school][teacher][guid]['student_count'] = 1
            else:
                sheet[school][teacher] = {guid: {id: {}}}
                sheet[school]['student_count'] += 1
                sheet[school][teacher]['student_count'] = 1
                sheet[school][teacher][guid]['student_count'] = 1
        else:
            sheet[school] = {teacher: {guid: {id: {}}}}
            sheet[school]['student_count'] = 1
            sheet[school][teacher]['student_count'] = 1
            sheet[school][teacher][guid]['student_count'] = 1
    
        duration = end_time - start_time
	if total == 0:
		print school, teacher, guid, id, session_name, finished, total
        percentage = float(finished)*100.0 / float(total)
    
        # add instance information to each student
        completed = sheet[school][teacher][guid][id].keys()
        enum_sessions = [re.sub('[\(\[].*?[\)\]]', '', x) for x in completed]
        if session_name in completed:
            session_name = session_name + '(' + str(enum_sessions.count(session_name)) + ')'
            sheet[school][teacher][guid][id][session_name] = [str(duration), finished,
                                                                 total, percentage]
        else:
            sheet[school][teacher][guid][id][session_name] = [str(duration), finished,
                                                         total, percentage]

    return sheet

if __name__ == "__main__":

    writer = ExcelWriter(session)

    ws1 = writer.workbook.active
    ws1.title = 'Stats_' + DATESTRING

    ws1.freeze_panes = 'E3'

    writer.create_title_bar(ws1)
    writer.add_log_data(ws1, log_dict=get_data())
    writer.add_legend()

    writer.save('Stats_' + DATESTRING + '.xlsx')

# dump data to JSON format
d = get_data()
with open('Stats_'+DATESTRING+'.json', 'w') as fp:
    json.dump(d, fp, sort_keys=True, indent=4, separators=[', ', ': '])
